import datetime
from decimal import Decimal, InvalidOperation
import requests
import locale
import threading
import json
import time
import base64
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import os

try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')

UF_API_URL = "https://mindicador.cl/api/uf"

UF_CACHE_FILE = os.path.join(os.path.dirname(__file__), 'uf_cache.json')

cache_lock = threading.Lock()

def _load_uf_from_file():
    
    try:
        with open(UF_CACHE_FILE, 'r') as f:
            data = json.load(f)
            return {
                "date": datetime.date.fromisoformat(data["date"]),
                "value": Decimal(data["value"])
            }
    except (FileNotFoundError, json.JSONDecodeError, KeyError):
        
        return {
            "date": datetime.date.today(),
            "value": Decimal('37000.00')
        }

def _save_uf_to_file(data):
   
    with open(UF_CACHE_FILE, 'w') as f:
        json.dump({"date": str(data["date"]), "value": str(data["value"])}, f)

def format_rut(rut_str):
    if not rut_str:
        return ''
    rut_str = str(rut_str).replace('.', '').replace('-', '').strip().upper()
    if len(rut_str) < 2:
        return rut_str
    
    body = rut_str[:-1]
    verifier = rut_str[-1]
    
    body_formatted = f"{int(body):,}".replace(",", ".")
    return f"{body_formatted}-{verifier}"

def _generate_excel_file(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Pagos"

    headers = ["Residente", "RUT", "Período/Detalle", "Fecha de Pago", "Monto Arriendo", "Monto Multa", "Total Pagado", "Estado", "Observaciones"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    currency_format = '$ #,##0'
    date_format = 'DD-MM-YYYY'
    
    for rec in records:
        is_adjustment = rec.get('estado') == 'Ajuste'
        periodo_display = ''
        if is_adjustment:
            periodo_display = (rec.get('observaciones') or 'Ajuste').split('\n')[0]
        elif rec.get('periodo'):
            try:
                periodo_date = datetime.datetime.strptime(rec['periodo'], '%Y-%m-%d')
                periodo_display = periodo_date.strftime('%B %Y').capitalize()
            except (ValueError, TypeError):
                periodo_display = rec['periodo']

        fecha_pago_val = None
        if rec.get('fecha_pago'):
            try:
                fecha_pago_val = datetime.datetime.strptime(rec['fecha_pago'].split(' ')[0], '%Y-%m-%d')
            except (ValueError, TypeError):
                fecha_pago_val = rec['fecha_pago']

        row_data = [
            rec.get('residente_nombre', ''),
            format_rut(rec.get('residente_rut', '')),
            periodo_display,
            fecha_pago_val,
            Decimal(rec.get('monto_arriendo', '0') or '0') if not is_adjustment else None,
            Decimal(rec.get('monto_multa', '0') or '0') if not is_adjustment else None, # Corregido: monto_pagado en lugar de total_pagado
            Decimal(rec.get('monto_pagado', '0') or '0'),
            rec.get('estado', ''),
            rec.get('observaciones', '')
        ]
        ws.append(row_data)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.cell(row=row_idx, column=4).number_format = date_format
        ws.cell(row=row_idx, column=5).number_format = currency_format
        ws.cell(row=row_idx, column=6).number_format = currency_format
        ws.cell(row=row_idx, column=7).number_format = currency_format
        
    column_widths = {'A': 30, 'B': 15, 'C': 25, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 12, 'I': 40}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    
    return base64.b64encode(virtual_workbook.getvalue()).decode('utf-8')

def _generate_pdf_file(records, summary):
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=0.25*inch, leftMargin=0.25*inch,
                            topMargin=1.5*inch, bottomMargin=0.5*inch)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Right', alignment=2))
    styles.add(ParagraphStyle(name='Center', alignment=1))
    styles.add(ParagraphStyle(name='PageInfo', alignment=1, fontSize=8, textColor=colors.grey))
    styles['Title'].fontSize = 18
    styles['Title'].spaceAfter = 0
    
    def format_clp(value_str):
        if value_str is None: return '$0'
        try:
            num = Decimal(value_str)
            return f"${int(num):,}".replace(",", ".")
        except (InvalidOperation, ValueError):
            return '$0'

    def header_footer(canvas, doc):
        canvas.saveState()
        width, height = doc.pagesize

       
        logo_path = os.path.join(os.path.dirname(__file__), '..', 'ui', 'templates', 'car_parking_logo.png')
        if os.path.exists(logo_path):
            logo_height = 1.25 * inch
            logo = Image(logo_path, width=1.25*inch, height=logo_height)
           
            logo_y_pos = height - doc.topMargin + (doc.topMargin - logo_height) / 2
            logo.drawOn(canvas, doc.leftMargin, logo_y_pos)

        header_text = Paragraph(
            "<b>Reporte de Historial de Pagos</b><br/>" +
            f"<font size=9>Generado el: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
            styles['Right']
        )
        header_text.wrap(width - doc.rightMargin - (doc.leftMargin + 1.5*inch), doc.topMargin)
        header_text.drawOn(canvas, doc.leftMargin + 1.5*inch, logo_y_pos)
        
        canvas.setStrokeColorRGB(0.2, 0.2, 0.2)
        
        canvas.line(doc.leftMargin, height - doc.topMargin, width - doc.rightMargin, height - doc.topMargin)

       
        page_num_text = f"Página {doc.page}"
        page_info = Paragraph(page_num_text, styles['PageInfo'])
        page_info.wrap(width, doc.bottomMargin)
        page_info.drawOn(canvas, 0, doc.bottomMargin - 0.3*inch)
        
        canvas.line(doc.leftMargin, doc.bottomMargin, width - doc.rightMargin, doc.bottomMargin)
        
        canvas.restoreState()

    
    summary_data = [
        [Paragraph('<b>Total Arriendos</b>', styles['Normal']), format_clp(summary.get('total_arriendo'))],
        [Paragraph('<b>Total Multas</b>', styles['Normal']), format_clp(summary.get('total_multas'))],
        [Paragraph('<b>Total Ajustes</b>', styles['Normal']), format_clp(summary.get('total_ajustes'))],
        [Paragraph('<b>TOTAL GENERAL</b>', styles['h4']), Paragraph(f"<b>{format_clp(summary.get('total_general'))}</b>", styles['h4'])]
    ]
    summary_table = Table(summary_data, colWidths=[1.5*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('LINEBELOW', (0, 2), (-1, 2), 0.5, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.25*inch))

    
    header = [Paragraph(h, styles['h5']) for h in ["Residente", "RUT", "Período", "Fecha Pago", "M. Arriendo", "M. Multa", "Total Pagado", "Estado"]]
    data = [header]

    for rec in records:
        is_adjustment = rec.get('estado') == 'Ajuste'
        periodo_display = ''
        if is_adjustment:
            periodo_display = (rec.get('observaciones') or 'Ajuste').split('\n')[0]
        elif rec.get('periodo'):
            periodo_date = datetime.datetime.strptime(rec['periodo'], '%Y-%m-%d')
            periodo_display = periodo_date.strftime('%B %Y').capitalize()

        fecha_pago_display = datetime.datetime.strptime(rec['fecha_pago'].split(' ')[0], '%Y-%m-%d').strftime('%d/%m/%Y') if rec.get('fecha_pago') else ''

        data.append([
            Paragraph(rec.get('residente_nombre', ''), styles['Normal']),
            Paragraph(format_rut(rec.get('residente_rut', '')), styles['Normal']),
            Paragraph(periodo_display, styles['Normal']),
            Paragraph(fecha_pago_display, styles['Center']),
            Paragraph(format_clp(rec.get('monto_arriendo')) if not is_adjustment else '-', styles['Right']),
            Paragraph(format_clp(rec.get('monto_multa')) if not is_adjustment else '-', styles['Right']),
            Paragraph(format_clp(rec.get('monto_pagado')), styles['Right']),
            Paragraph(rec.get('estado', ''), styles['Center']),
        ])

    table = Table(data, colWidths=[1.6*inch, 1.0*inch, 1.0*inch, 0.9*inch, 0.9*inch, 0.8*inch, 0.9*inch, 0.67*inch], splitByRow=1, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 1), (2, -1), 'LEFT'),
        ('ALIGN', (4, 1), (-2, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')

def _generate_csv_file(records):
    output = StringIO()
    writer = csv.writer(output)
    
    headers = ["ID", "Residente", "RUT", "Periodo", "Fecha Pago", "Monto Arriendo", "Monto Multa", "Total Pagado", "Observaciones", "Estado"]
    writer.writerow(headers)

    for rec in records:
        periodo_display = ''
        if rec.get('periodo'):
            try:
                periodo_date = datetime.datetime.strptime(rec['periodo'], '%Y-%m-%d')
                periodo_display = periodo_date.strftime('%B %Y').capitalize()
            except (ValueError, TypeError):
                periodo_display = rec['periodo']
        
        fecha_pago_display = ''
        if rec.get('fecha_pago'):
            fecha_pago_display = rec['fecha_pago'].split(' ')[0]

        row = [
            rec.get('id', ''),
            rec.get('residente_nombre', ''),
            format_rut(rec.get('residente_rut', '')),
            periodo_display,
            fecha_pago_display,
            rec.get('monto_arriendo', '0') or '0', 
            rec.get('monto_multa', '0') or '0',
            rec.get('monto_pagado', '0') or '0',
            rec.get('observaciones', ''),
            rec.get('estado', '')
        ]
        writer.writerow(row)
        
    return base64.b64encode(output.getvalue().encode('utf-8')).decode('utf-8')

def update_uf_cache():
    global uf_data_cache
    try:
        response = requests.get(UF_API_URL, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        uf_entry = data['serie'][0]
        uf_value = Decimal(str(uf_entry['valor']))
        date_str = uf_entry['fecha'].split('T')[0]
        uf_date = datetime.date.fromisoformat(date_str)

        with cache_lock:
            uf_data_cache["date"] = uf_date
            uf_data_cache["value"] = uf_value
            _save_uf_to_file(uf_data_cache) 
        
        print(f"✅ UF actualizada desde mindicador.cl: {uf_value} para la fecha {uf_date.strftime('%d/%m/%Y')}")
            
    except (requests.exceptions.RequestException, Exception) as e:
        print(f"❌ No se pudo actualizar la UF desde la API de mindicador.cl. Se usará el último valor conocido. Error: {e}")


def periodic_uf_updater():
    while True:
        update_uf_cache()
        time.sleep(4 * 60 * 60)

def get_uf_data():
    with cache_lock:
        return {
            'uf_value': str(uf_data_cache['value']),
            'current_date': uf_data_cache['date'].strftime('%d/%m/%Y')
        }

def get_all_active_residents_for_dropdown(db_manager):
    return db_manager.get_all_active_residents_for_dropdown()

def create_payment_adjustment(db_manager, resident_id, periodo, monto, observaciones):
    return db_manager.create_payment_adjustment(resident_id, periodo, monto, observaciones)

def delete_payment_record(db_manager, payment_id):
    return db_manager.delete_payment_record(payment_id)

def get_payment_history(db_manager, filters):
    return db_manager.get_payment_history(filters)

def update_payment_record(db_manager, payment_id, data):
    return db_manager.update_payment_record(payment_id, data)

def get_resident_status_list(db_manager):
    with cache_lock:
        current_date = uf_data_cache['date']
        uf_value = uf_data_cache['value']

    all_residents_contracts = db_manager.get_all_active_residents_for_status()
    resident_payments = db_manager.get_all_payment_records()

    payments_by_contract = {}
    for payment in resident_payments:
        contract_id = payment['id_contrato']
        if contract_id not in payments_by_contract:
            payments_by_contract[contract_id] = set()
        payments_by_contract[contract_id].add(payment['periodo'])

    status_list = []
    
  
    residents_data = {}
    for contract in all_residents_contracts:
        res_id = contract['id_residente']
        if res_id not in residents_data:
            residents_data[res_id] = {
                "id_residente": res_id,
                "nombre_completo": contract['nombre_completo'],
                "rut": contract['rut'],
                "contratos": []
            }
        residents_data[res_id]['contratos'].append(contract)

    for resident_id, resident in residents_data.items():
        status = "Al día"
        dias_de_atraso = 0
        dia_pago_display = set()
        all_pagos_realizados = set()
        tiene_pagos_adelantados = False

        for contrato in resident['contratos']:
            fecha_inicio = datetime.date.fromisoformat(contrato['fecha_inicio']) if isinstance(contrato['fecha_inicio'], str) else contrato['fecha_inicio']
            dia_pago_display.add(str(fecha_inicio.day))
            pagos_del_contrato = payments_by_contract.get(contrato['id_contrato'], set())
            all_pagos_realizados.update(pagos_del_contrato)

            contrato_status = "Al día"
            mes_a_evaluar = fecha_inicio.replace(day=1)
            while mes_a_evaluar <= current_date.replace(day=1):
                periodo_str = mes_a_evaluar.strftime('%Y-%m-%d')
                if periodo_str not in pagos_del_contrato:
                    fecha_vencimiento = mes_a_evaluar.replace(day=fecha_inicio.day)
                    if current_date > fecha_vencimiento:
                        
                        contrato_status = "Pendiente" 
                        atraso_actual = (current_date - fecha_vencimiento).days
                        dias_de_atraso = max(dias_de_atraso, atraso_actual)
                        if dias_de_atraso >= 30:
                            contrato_status = "Atrasado"
                        break 
                
                mes_a_evaluar = (mes_a_evaluar.replace(day=28) + datetime.timedelta(days=4)).replace(day=1)

            if (status == "Al día" and contrato_status in ["Pendiente", "Atrasado"]) or \
               (status == "Pendiente" and contrato_status == "Atrasado"):
                status = contrato_status

        if status == "Al día":
            mes_siguiente = (current_date.replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
            if mes_siguiente.strftime('%Y-%m-%d') in all_pagos_realizados:
                tiene_pagos_adelantados = True

        final_status = status
        if status == "Al día" and tiene_pagos_adelantados:
            final_status = "Adelantado"
        elif status == "Atrasado" and dias_de_atraso < 30:
            final_status = "Pendiente"

        status_list.append({
            "id_residente": resident['id_residente'],
            "nombre_completo": resident['nombre_completo'],
            "rut": resident['rut'],
            "nombre_contrato": ", ".join([c.get('nombre_contrato', 'N/A') for c in resident['contratos']]),
            "estado": final_status,
            "dia_pago": ", ".join(sorted(list(dia_pago_display))),
            "dias_atraso": dias_de_atraso
        })
    
    return {
        'status_list': status_list,
        'uf_value': str(uf_value),
        'current_date': current_date.strftime('%d/%m/%Y')
    }

def get_resident_debt_details(db_manager, resident_id):
    with cache_lock:
        current_date = uf_data_cache['date']
        uf_value = uf_data_cache['value']
        
    details = db_manager.get_resident_contract_details(resident_id)
    if not details:
        return {"error": "Residente no encontrado"}

    pagos_realizados_raw = db_manager.get_payments_by_resident(resident_id)
    pagos_realizados = {p['periodo'] for p in pagos_realizados_raw}

    monto_mensual_por_contrato = {details['id_contrato']: calculate_expected_fee(details)} if details else {}
    
    meses_adeudados = []
    meses_futuros_disponibles = []

    if details and isinstance(details['fecha_inicio'], str):
        fecha_inicio = datetime.date.fromisoformat(details['fecha_inicio'])
    else:
        fecha_inicio = details['fecha_inicio']

    dia_pago = fecha_inicio.day
    mes_a_evaluar = fecha_inicio.replace(day=1)
    
    fecha_limite_futura = (current_date + datetime.timedelta(days=540)).replace(day=1)

    while mes_a_evaluar < fecha_limite_futura:
        periodo_str = mes_a_evaluar.strftime('%Y-%m-%d')
        if periodo_str not in [p['periodo'] for p in pagos_realizados_raw if p['estado'] in ['Pagado', 'Ajuste']]:
            mes_info = {
                "periodo_display": mes_a_evaluar.strftime('%B %Y').capitalize(),
                "periodo_iso": periodo_str,
                "monto": str(monto_mensual_por_contrato.get(details['id_contrato'], Decimal('0')))
            }
            if mes_a_evaluar <= current_date.replace(day=1):
                meses_adeudados.append(mes_info)
            else:
                meses_futuros_disponibles.append(mes_info)

        if mes_a_evaluar.month == 12:
            mes_a_evaluar = mes_a_evaluar.replace(year=mes_a_evaluar.year + 1, month=1)
        else:
            mes_a_evaluar = mes_a_evaluar.replace(month=mes_a_evaluar.month + 1)
    
    multas_info = {"cantidad": 0, "monto_total": "0.00", "monto_unitario": "0.00"}
    if meses_adeudados:
        primer_mes_adeudado_date = datetime.date.fromisoformat(meses_adeudados[0]['periodo_iso'])
        fecha_vencimiento = primer_mes_adeudado_date.replace(day=dia_pago)
        
        if current_date > fecha_vencimiento:
            dias_de_atraso = (current_date - fecha_vencimiento).days
            cantidad_multas = dias_de_atraso // 30

            if cantidad_multas > 0:
                precio_multa_uf_str = details.get('precio_multa_uf')
                monto_multa_uf = Decimal(precio_multa_uf_str) if precio_multa_uf_str else Decimal('0.00')
                monto_multa_clp = (monto_multa_uf * uf_value).quantize(Decimal('0'))
                
                multas_info = {
                    "cantidad": cantidad_multas,
                    "monto_total": str(monto_multa_clp * cantidad_multas),
                    "monto_unitario": str(monto_multa_clp)
                }

    return {
        "nombre_completo": details['nombre_completo'],
        "meses_adeudados": sorted(meses_adeudados, key=lambda x: x['periodo_iso']),
        "meses_futuros_disponibles": sorted(meses_futuros_disponibles, key=lambda x: x['periodo_iso'])[:12],
        "multas": multas_info
    }

def process_payment(db_manager, resident_id, meses_a_pagar, cobrar_multas):
    details = db_manager.get_resident_contract_details(resident_id)
    if not details:
        return False, "No se encontraron los detalles del residente."

    id_contrato = details['id_contrato']
    
    pagos_a_registrar = []
    
    for mes in meses_a_pagar:
        pagos_a_registrar.append({
            "id_contrato": id_contrato,
            "periodo": mes['periodo_iso'],
            "monto_esperado": Decimal(mes['monto']),
            "monto_multa": Decimal('0.00'),
            "monto_pagado": Decimal(mes['monto']),
            "observaciones": "Pago de arriendo mensual."
        })

    if cobrar_multas.get('aplicar') and meses_a_pagar:
        pagos_a_registrar.append({
            "id_contrato": id_contrato,
            "periodo": meses_a_pagar[0]['periodo_iso'],
            "monto_esperado": Decimal('0.00'),
            "monto_multa": Decimal(cobrar_multas.get('monto_total', '0')),
            "monto_pagado": Decimal(cobrar_multas.get('monto_total', '0')),
            "observaciones": f"Pago de {cobrar_multas.get('cantidad', 0)} multa(s) por atraso."
        })

    return db_manager.register_bulk_payments(pagos_a_registrar)

def calculate_expected_fee(contract_details):
    try:
        p1_auto = Decimal(contract_details.get('precio_primer_estacionamiento_auto') or '0')
        p2_auto = Decimal(contract_details.get('precio_segundo_estacionamiento_auto') or '0')
        p1_moto = Decimal(contract_details.get('precio_estacionamiento_moto') or '0')
        p2_moto = Decimal(contract_details.get('precio_segundo_estacionamiento_moto') or '0')
    except (InvalidOperation, TypeError):
        return Decimal(0)

    total_fee = Decimal(0)
    count_auto = contract_details.get('autos_count', 0)
    count_moto = contract_details.get('motos_count', 0)

    if p2_auto > 0 and p2_auto is not None:
        num_pairs_auto = count_auto // 2
        remaining_auto = count_auto % 2
        total_fee += num_pairs_auto * (p1_auto + p2_auto)
        if remaining_auto == 1:
            total_fee += p1_auto
    else:
        total_fee += count_auto * p1_auto

    if p2_moto > 0 and p2_moto is not None:
        num_pairs_moto = count_moto // 2
        remaining_moto = count_moto % 2
        total_fee += num_pairs_moto * (p1_moto + p2_moto)
        if remaining_moto == 1:
            total_fee += p1_moto
    else:
        total_fee += count_moto * p1_moto

    return total_fee

def export_payment_history_to_excel(records):
    if not records:
        return None
    return _generate_excel_file(records)

def export_payment_history_to_csv(records):
    if not records:
        return None
    return _generate_csv_file(records)

def export_full_history_to_excel(db_manager, filters):
    history_data = db_manager.get_payment_history(filters)
    if history_data.get("error") or not history_data.get("records"):
        return None
    return _generate_excel_file(history_data["records"])

def export_full_history_to_csv(db_manager, filters):
    history_data = db_manager.get_payment_history(filters)
    if history_data.get("error") or not history_data.get("records"):
        return None
    return _generate_csv_file(history_data["records"])

def export_payment_history_to_pdf(db_manager, filters):
    history_data = db_manager.get_payment_history(filters)
    if history_data.get("error") or not history_data.get("records"):
        return None
    return _generate_pdf_file(history_data["records"], history_data["summary"])

def export_payment_history_to_pdf_current_view(records):
    if not records:
        return None
    
    total_arriendo = sum(Decimal(r.get('monto_arriendo', '0') or '0') for r in records if r['estado'] != 'Ajuste')
    total_multas = sum(Decimal(r.get('monto_multa', '0') or '0') for r in records if r['estado'] != 'Ajuste')
    total_ajustes = sum(Decimal(r.get('monto_pagado', '0') or '0') for r in records if r['estado'] == 'Ajuste')
    total_general = sum(Decimal(r.get('monto_pagado', '0') or '0') for r in records)

    summary = {"total_arriendo": str(total_arriendo), "total_multas": str(total_multas),
               "total_ajustes": str(total_ajustes), "total_general": str(total_general)}
    return _generate_pdf_file(records, summary)

def export_audit_log_to_excel(db_manager):
    records = db_manager.get_payment_audit_log()
    if not records:
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría de Pagos"
    headers = ["ID Historial", "ID Registro Pago", "Acción", "Fecha Acción", "ID Contrato", "Período", "Fecha Pago", "Monto Esperado", "Monto Multa", "Estado", "Monto Pagado", "Observaciones"]
    ws.append(headers) # type: ignore

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for rec in records:
        ws.append([ # type: ignore
            rec.get('id_historial'), rec.get('id_registro_pago'), rec.get('accion'),
            rec.get('fecha_accion'), rec.get('id_contrato'), rec.get('periodo'),
            rec.get('fecha_pago'), rec.get('monto_esperado'), rec.get('monto_multa'),
            rec.get('estado'), rec.get('monto_pagado'), rec.get('observaciones')
        ])
        
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].best_fit = True # type: ignore

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    return base64.b64encode(virtual_workbook.getvalue()).decode('utf-8')

def export_audit_log_to_csv(db_manager):
    records = db_manager.get_payment_audit_log()
    if not records:
        return None
        
    output = StringIO()
    writer = csv.writer(output)
    headers = ["id_historial", "id_registro_pago", "accion", "fecha_accion", "id_contrato", "periodo", "fecha_pago", "monto_esperado", "monto_multa", "estado", "monto_pagado", "observaciones"]
    writer.writerow(headers)
    
    for rec in records:
        writer.writerow([rec.get(h) for h in headers])
        
    return base64.b64encode(output.getvalue().encode('utf-8')).decode('utf-8')


uf_data_cache = _load_uf_from_file()


update_uf_cache()